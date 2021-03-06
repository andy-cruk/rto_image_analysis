'''Script will collect users for each segment, all segments for a core, and expert scores for those cores. Stores results in excel and mongodb.

Code can also calculate Spearman correlations between expert and citizen science for varying number of citizen scientists/segment. However,
the method used here, which aggregates all classifications for each subject (=segment), is inefficient. So this is an experimental
feature that should not be used for heavy lifting.

This script generates a number of dataframes in order
cln         classifications: one row per subject (aggregated over users, or not aggregated at all)
cores       one row per core (aggregated over subjects)
'''
__author__ = 'Peter'
from pymongo import MongoClient
import pandas as pd
import numpy as np
import os.path
import os
import matplotlib.pyplot as plt
from sklearn import linear_model
from sklearn import cross_validation as cv
from sklearn import pipeline
from sklearn.preprocessing import Imputer
import time
import re
from openpyxl import load_workbook
import datetime
import quadratic_weighted_kappa as qwk
import rto_mongodb_utils
import math

pd.set_option('display.max_columns', 500)
pd.set_option('expand_frame_repr', False)

# USER OPTIONS
# currently done (feb 12 2016): mre11, p21, 53bp1, p53, rad50, ck5, mre11new, ck20, tip60, ki67, mre11c
stain = "ki67".lower()  # what sample to look at; must match metadata.stain_type_lower in subjects database,e.g. "TEST MRE11" or "MRE11", "rad50", "p21". Case-INSENSITIVE because the database is queried for upper and lower case version
aggregate = 'ignoring_segments'      # how to aggregate, also field that is written to in mongodb. 'ignoring_segments' or 'segment_aggregation'
# aggregate = 'segment_aggregation'
bootstrap = True       # whether to bootstrap
print(stain, aggregate)

if (aggregate == "ignoring_segments") & (not bootstrap):
    numberOfClassificationsPerCore = np.array([0])  # will draw X classifications per core with replacement. Only used if aggregate = 'ignoring_segments'. Set to zero to include all; give a range to test multiple numbers
    samplesPerNumberOfUsers = 1       # for each value in numberOfUsersPerSubject, how many times to sample users with replacement. Set to 1 if you just want to run once, e.g. when you include all the users
elif (aggregate == "ignoring_segments") & bootstrap:
    numberOfClassificationsPerCore = np.array([1,2,4,8,16,32,64,128,256,512,1024])
    samplesPerNumberOfUsers = 1000       # for each value in numberOfUsersPerSubject, how many times to sample users with replacement. Set to 1 if you just want to run once, e.g. when you include all the users
elif (aggregate == "segment_aggregation") & (not bootstrap):
    numberOfUsersPerSubject = np.array([0]) # will loop over each of the number of users and calculate Spearman rho. Only used if aggregate = 'segment_aggregation'. Set to 0 to not restrict number of users
    samplesPerNumberOfUsers = 1       # for each value in numberOfUsersPerSubject, how many times to sample users with replacement. Set to 1 if you just want to run once, e.g. when you include all the users
elif (aggregate == "segment_aggregation") & bootstrap:
    numberOfUsersPerSubject = np.array([1,2,3,4,5,6]) # will loop over each of the number of users and calculate Spearman rho. Only used if aggregate = 'segment_aggregation'. Set to 0 to not restrict number of users
    samplesPerNumberOfUsers = 1000       # for each value in numberOfUsersPerSubject, how many times to sample users with replacement. Set to 1 if you just want to run once, e.g. when you include all the users

# set dictionary with filters to feed to mongoDB. If lowercase versions don't exist use rto_mongodb_utils to add lowercase versions
filterSubjects = {"$and": [
    {"metadata.stain_type_lower": stain},
    {"classification_count": {"$gte": 1}},
]}

# list all stain types we have GS for in a list
f = [fn for fn in os.listdir('GS') if ('GS_' in fn) and ('.xlsx' in fn)]
stains = [x.lstrip('GS_').rstrip('.xlsx') for x in f]

# save file for scores
classificationsDataframeFn = "results\classifications_dataframe_" + stain + ".pkl"  # will store the pandas dataframe created from all the classifications; file will load if the file exists already to prevent another 30 to 60 min of processing.


########### FUNCTION DEFINITIONS
def pymongo_connection_open():
    """ open connection to local data and return subjectsCollection, classifCollection, and dbConnection
    Change as appropriate if your database has a different name.
    """
    dbConnection = MongoClient("localhost", 27017)
    subjectsCollection = dbConnection[rto_mongodb_utils.currentDB].subjects
    classifCollection = dbConnection[rto_mongodb_utils.currentDB].classifications
    # nSj = subjectsCollection.count()
    # nCl = classifCollection.count()
    return subjectsCollection, classifCollection, dbConnection
def pymongo_connection_close():
    """Close pymongo connection"""
    dbConnection.close()
def classifications_dataframe_fill(numberOfUsersPerSubject,skipNonExpertClassifications=False):
    """Find all classifications for requested stain and aggregate responses for each subject
    This function saves and returns a pandas dataframe that has one row per subject (=segment) and columns that indicate proportion of all classifications given that answer.
    Note: proportion for different stain answers might not add up to 1 as people that say 'no cancer' still count towards classifications
    Note that you should have set an index in your mongoDB in the db.classifications.subject_ids field, otherwise this code will be agonisingly slow.
    """
    # first grab all the subjects that satisfy the filterSubjects criteria. Exclude those that do not have expert if requested
    filterCopy = filterSubjects.copy()
    if skipNonExpertClassifications:
        filterCopy["hasExpert"] = True
    subjectCursor = subjectsCollection.find(filter=filterCopy, projection=("classification_count", "metadata.id_no"),no_cursor_timeout=True)

    assert subjectCursor.count() > 0, "no records found for filter"
    col = ("subjectID", "core", "nClassifications", "nCancer", "0%Stain", "<25%Stain", "<50%Stain", "<75%Stain", "<95%Stain",
           "<100%Stain", "stainNone", "stainWeak", "stainMedium", "stainStrong")
    # set up dataframe that will store one row per segment, with classification scores. It will count occurrences in each of the buckets, which can later be divided by number of classifications
    # Trying to initialise with sufficient rowsbased on timeit.timeit("df.loc[ix,:]=np.nan;ix+=1",setup='''import pandas as pd; import numpy as np; df = pd.DataFrame(data=np.zeros((1000,14)),columns=("subjectID", "core", "nClassifications", "nCancer", "0%Stain", "<25%Stain", "<50%Stain", "<75%Stain", "<95%Stain","<100%Stain", "stainNone", "stainWeak", "stainMedium", "stainStrong"));ix=0''',number=1000)
    cl = pd.DataFrame(data=None, columns=col)
    clIx = 0  # dataframe index
    for sj in subjectCursor:  # loop over each entry in subjectCursor and aggregate ratings into dataframe
        # store the subject id in the pandas dataframe, initialising the row
        cl.loc[clIx, "subjectID"] = sj["_id"]
        # store what core it comes from
        cl.loc[clIx, "core"] = sj["metadata"]["id_no"]
        # initialise the rest of the row with zeros
        cl.iloc[clIx, 2:] = 0
        # collect all classifications from this particular segment and put in a numpy array for easy indexing later
        clCursor = np.array(list(classifCollection.find(filter={"subject_ids": [sj["_id"]]},projection=("annotations",))))
        # select a random subset of items to be selected from the cursor, or select all if all users were requested or not sufficient users available
        if numberOfUsersPerSubject == 0: # all requested
            selectedSubjects = np.arange(len(clCursor))
        elif (numberOfUsersPerSubject > 0) and (numberOfUsersPerSubject <= len(clCursor)): # if number of users requested between 0 and max number of users for this subject
            selectedSubjects = np.random.choice(len(clCursor),numberOfUsersPerSubject,replace=False)
        elif (numberOfUsersPerSubject > 0) and (numberOfUsersPerSubject > len(clCursor)): # if for this segment not enough users are available, select all
            selectedSubjects = np.arange(len(clCursor))
        for iCl in clCursor[selectedSubjects]:  # loop over each classification in selectedSubjects, which is either all classifications or a randomly selected subset
            cl.loc[clIx, "nClassifications"] += 1
            # cancer yes/no store in dataframe
            cancer = int(iCl["annotations"][0]["a-1"])
            if cancer == 1:  # if yes cancer
                propStain = int(iCl["annotations"][1]["a-2"])           #retrieve proportion stained category
                intensityStain = int(iCl["annotations"][2]["a-3"])      #retrieve intensity staining category
                # check if this is an invalid entry because proportion should not be 0 (given there's cancer, 1 would indicate no staining). Similarly, if propStain > 1 then intensityStain should be >0
                if (propStain == 0) or (propStain>1 and intensityStain==0):
                    # reduce number of classifications by 1, effectively taking it out of the tally
                    print (cancer,propStain,intensityStain)
                    cl.loc[clIx, "nClassifications"] -= 1
                    # continue the loop with next classification
                    continue
                cl.loc[clIx, "nCancer"] += 1
                if propStain == 1:
                    cl.loc[clIx, "0%Stain"] += 1
                elif propStain == 2:
                    cl.loc[clIx, "<25%Stain"] += 1
                elif propStain == 3:
                    cl.loc[clIx, "<50%Stain"] += 1
                elif propStain == 4:
                    cl.loc[clIx, "<75%Stain"] += 1
                elif propStain == 5:
                    cl.loc[clIx, "<95%Stain"] += 1
                elif propStain == 6:
                    cl.loc[clIx, "<100%Stain"] += 1
                # now do intensity
                if intensityStain == 0:
                    cl.loc[clIx, "stainNone"] += 1
                elif intensityStain == 1:
                    cl.loc[clIx, "stainWeak"] += 1
                elif intensityStain == 2:
                    cl.loc[clIx, "stainMedium"] += 1
                elif intensityStain == 3:
                    cl.loc[clIx, "stainStrong"] += 1
            elif cancer == 2:  # if no cancer
                continue

        clIx += 1
    # convert numbers to int (otherwise they're stored as object)
    cl = cl.convert_objects(convert_numeric=True)
    # normalise everything but nClassifications to value between 0 and 1
    # create cln, which is normalised version of cl
    cln = cl.copy()
    # normalise each of the columns starting at nCancer and all the ones to the right, using my own function def
    cln.loc[:,"nCancer":] = normalise_dataframe_by_ix(cl,cl.columns.get_loc("nClassifications"),range(cl.columns.get_loc("nCancer"),len(cl.columns)))
    return cln
def classifications_dataframe_fill_individual_classifications(skipNonExpertClassifications=False):
    """ This will look at the requested stain and grab all the classifications that correspond to it.
    Returns this as a dataframe with sensible names for columns

    :param skipNonExpertClassifications: bool indicating whether to only select classifications that were performed on a core with expert info
    :return: a pandas dataframe, nClassifications*nColumns
    """
    # retrieve classifications, either including or excluding those on non-expert cores
    project = {
            '_id': False,
            'cancer': True,
            'proportion': True,
            'intensity': True,
            'id_no': True}
    if skipNonExpertClassifications:
        clCursor = classifCollection.find({"stain_type_lower":stain, "hasExpert":True}, projection=project)
    if ~skipNonExpertClassifications:
        clCursor = classifCollection.find({"stain_type_lower":stain}, projection=project)
    # check something was actually found
    assert clCursor.count()>0
    # dump data into dataframe. This can take a long time and requires a lot of memory
    cln = pd.DataFrame(list(clCursor))
    # set cancer to True or False
    cln.cancer = cln.cancer.apply(int) == 1
    # set noCancer classifications IHC scores to 0
    cln.loc[~cln.cancer,'intensity'] = np.nan
    cln.loc[~cln.cancer,'proportion'] = np.nan
    cln = cln.convert_objects(convert_numeric=True)
    # add percentage stained
    cln['aggregatePropWeighted'] = category_to_percentage(cln.proportion-1)
    # rename columns to sensible descriptions that rest of the code understands
    cln.rename(columns={
        'id_no': 'core',
        "_id": "subjectID",
        'proportion': 'aggregatePropWeightedCategory',
        'intensity': 'aggregateIntensityWeighted'}
        ,inplace=True)

    return cln
def sj_in_expert_core(coreID,coresGS):
    """
    Checks for a given core object ID whether it can be found in the expert scores. Used by classifications_dataframe_fill
    to save having to fill the cln dataframe with segments that do not belong to a core we have GS for.
    :param id: a core id that will be checked against coresGS
    :return: True or False
    """
    for ixGS,rowGS in coresGS.iterrows():
        # if the current row in coresGS is the one that matches the coreID, return True
        if str(int(rowGS.loc["coreID"])) in coreID:
            return True
    # if you get to this part of the code, no match has been found in rowGS
    return False
def classifications_dataframe_save(cln,fn=classificationsDataframeFn):
    print("Saving dataframe to %s" % fn)
    cln.to_pickle(fn)
def classifications_dataframe_load(fn=classificationsDataframeFn):
    """load dataframe and return pandas dataframe"""
    cln = pd.read_pickle(fn)
    return cln
def plot_weighted_vs_unweighted_stain(cores):
    """Plots scores for aggregate staining intensity and proportion for weighted vs unweighted instance; weighting happens across subjects based on their proportion of people who said it contained cancer
    """
    plt.figure(1)
    plt.scatter(cores.aggregateIntensity,cores.aggregateIntensityWeighted)
    plt.xlabel("unweighted")
    plt.ylabel("weighted")
    plt.title("Aggregate Intensity")
    plt.figure(2)
    plt.scatter(cores.aggregateProp,cores.aggregatePropWeighted)
    plt.xlabel("unweighted")
    plt.ylabel("weighted")
    plt.title("Aggregate Proportion")
    plt.show()
def plot_user_vs_expert(cores):
    """Plot intensity, proportion and combined scores for experts against users, and print spearman
    :param cores: pandas dataframe
    :return: None
    """
    rhoProp,rhoIntensity,rhoSQS,rhoSQSadditive = user_vs_expert_rho(cores)
    plt.subplot(3,1,1)
    plt.scatter(cores.expProp,cores.aggregatePropWeighted)
    plt.xlabel("expert")
    plt.ylabel("user")
    plt.title("proportion stained, Spearman r = "+'{0:.2f}'.format(rhoProp))
    plt.gca().set_aspect('equal', adjustable='box')
    plt.tight_layout(2)

    plt.subplot(3,1,2)
    plt.scatter(cores.expIntensity,cores.aggregateIntensityWeighted)
    plt.xlabel("expert")
    plt.ylabel("user")
    plt.title("intensity stain, Spearman r = "+'{0:.2f}'.format(rhoIntensity))
    plt.gca().set_aspect('equal', adjustable='box')

    plt.subplot(3,1,3)
    plt.scatter(cores.expSQS,cores.aggregateSQS)
    plt.xlabel("expert")
    plt.ylabel("user")
    plt.title("combined scores, Spearman r = "+'{0:.2f}'.format(rhoSQS))
    plt.gca().set_aspect('equal', adjustable='box')

    fig = plt.gcf()
    fig.canvas.set_window_title(stain)

    plt.show()
def cln_add_columns_aggregating_stain(cln):
    """Add 2 columns that aggregate stain proportion and intensity answers into new columns
    Will disregard those that answered 'not cancer' and calculate aggregates based on only those people that said there is cancer.
    Stores NaN for any segment where no one said cancer
    Aggregate is calculated by
    Proportion: transforming to numbers, taking mean, transforming back to categories
    Intensity: mean across categories {1,2,3}
    Returns cln with 2 additional columns
    """
    _,intensities = percentage_to_category(None)
    # this next one's a beast and should probably be spread out. The numerator takes the n * 5 matrix of stain proportions and dot multiplies it with the mean intensities; effectively
    # summing the products of proportion*intensities. This is then divided by the proportion of people saying there was indeed cancer to correct
    # for the fact that we only want to include people that answered 'yes cancer'.
    cln["aggregateProp"] = (cln.loc[:,"0%Stain":"<100%Stain"].dot(np.array(intensities["percentage"]))) / cln.nCancer
    # same deal for intensity, but now we multiple simply by 1,2,3 for weak, medium, strong respectively.
    cln["aggregateIntensity"] = (cln.loc[:,"stainNone":"stainStrong"].dot(np.array([0,1,2,3]))) / cln.nCancer
    # For samples w/o anyone saying 'cancer', store a NaN
    cln.loc[cln["aggregateProp"]<0.0001,('aggregateProp','aggregateIntensity')] = np.nan
    return cln
def core_dataframe_fill(cln):
    """takes a dataframe cln (row=subject,column=properties of subject such as responses); aggregates all subjects for a single core into a single row in dataframe "cores"
    A subject with 20 classifications is weighted equally to subject with 150 classifications.
    For combining aggregratePropWeighted and aggregateIntensityWeighted, the contributions from different subjects are weighted by the probability of users saying
    the subject was cancer. For example, a subject with nCancer = 0.3 will be weighted at half that of a subject with nCancer = 0.6

    Some cores may not have segments in cln because cln only contains segments that belong to cores that we have GS for. These cores
    will have nans as data.
    """
    # initialise new DataFrame that will have one row per core AND THE SAME COLUMNS AS classifications
    cores = pd.DataFrame(data=None,index=range(0,len(get_core_ids(cln))),columns=cln.columns,dtype="float64")
    # put in core IDs
    cores["core"]=get_core_ids(cln)
    # get rid of unnecessary subjectID column
    cores = cores.drop("subjectID",axis=1)
    # add column indicating number of subjects for each core
    cores.insert(1,"nSubjects",np.nan)
    # add columns for weighted aggregates
    cores.insert(len(cores.columns),"aggregatePropWeighted",np.nan)
    cores.insert(len(cores.columns),"aggregateIntensityWeighted",np.nan)
    cores.insert(len(cores.columns),"aggregateSQS",np.nan)
    cores.insert(len(cores.columns),"aggregateSQSadditive",np.nan)
    # loop over each core
    for ix,core in cores.iterrows():
        coreRowsInCln = cln[cln["core"]==core.core]
        # note this takes the mean across all segments, including those that probably didn't have cancer.
        cores.loc[ix,"nClassifications":"aggregateIntensity"] = coreRowsInCln.loc[:,"nClassifications":"aggregateIntensity"].mean()
        # if none of the subjects have anyone saying there's cancer, set IHC to 0. Otherwise you get nans and inf which will mess you up later (e.g. when correcting scores)
        if coreRowsInCln.nCancer.sum() < np.finfo(float).eps: # should not do ==0.0 for float, so test for smaller than epsilon
            # add weighted aggregate scores; multiply each subject score by nCancer, then normalise by nCancer.sum()
            cores.loc[ix,("aggregatePropWeighted","aggregateIntensityWeighted","aggregateSQS","aggregateSQSadditive")] = np.nan
        else:
            # add weighted aggregate scores; multiply each subject score by nCancer, then normalise by nCancer.sum()
            cores.loc[ix,"aggregatePropWeighted"]       = (coreRowsInCln.aggregateProp      *coreRowsInCln.nCancer).sum() / coreRowsInCln.nCancer.sum()
            cores.loc[ix,"aggregateIntensityWeighted"]  = (coreRowsInCln.aggregateIntensity *coreRowsInCln.nCancer).sum() / coreRowsInCln.nCancer.sum()
            cores.loc[ix,"aggregateSQS"] = cores.loc[ix,"aggregatePropWeighted"]*cores.loc[ix,"aggregateIntensityWeighted"]
            cores.loc[ix,"aggregateSQSadditive"] = percentage_to_category([cores.loc[ix,"aggregatePropWeighted"]])[0] + cores.loc[ix,"aggregateIntensityWeighted"]
        # store how many subjects were included for the core
        cores.loc[ix,"nSubjects"] = len(coreRowsInCln.index)

    # add category for aggregateWeighted
    cores.insert(loc=cores.columns.get_loc("aggregatePropWeighted")+1,column="aggregatePropWeightedCategory",value=percentage_to_category(cores.aggregatePropWeighted)[0])
    return cores

def cores_dataframe_fill_from_individual_classifications(cln, nCl=0):
    """
    Takes a dataframe with classifications from classifications_dataframe_fill_individual_classifications() and
    calculates scores for each core
    :param cln: dataframe, see whatever is returned by classifications_dataframe_fill_individual_classifications()
    :param nCl: number of classifications to include per core; set to 0 to include all. Cannot request more than are available
    :return: df with nCores * {'aggregatePropWeighted','aggregateIntensityWeighted','aggregateSQS','aggregateSQSadditive','aggregatePropWeightedCategory'}
    """
    # get numpy array of core ID numbers, splitting off the stain type
    coreID = cln.core.unique()
    # set up function to randomly select classifications.
    # Thanks a million to http://stackoverflow.com/questions/22472213/python-random-selection-per-group
    if nCl == 0: # select all; will skip nan when taking mean
        fn = lambda obj: obj.mean()
    else: # select random set with replacement and take mean; will skip nan
        fn = lambda obj: obj.loc[np.random.choice(obj.index, nCl, True),:].mean()
    # apply the function and get the cores dataframe
    cores = cln.groupby('core').apply(fn)
    # store how many classifications were available for each core
    fn = lambda obj: len(obj)
    nClassificationsTotal = cln.groupby('core').apply(fn).to_frame('nClassificationsTotal')
    cores = cores.join(nClassificationsTotal, how='left')
    cores['aggregateSQS'] = cores.aggregatePropWeighted * cores.aggregateIntensityWeighted
    cores['aggregateSQSadditive'] = percentage_to_category(cores.aggregatePropWeighted)[0] + cores.aggregateIntensityWeighted
    cores['aggregatePropWeightedCategory'] = percentage_to_category(cores.aggregatePropWeighted)[0]
    # put core column back into dataframe
    cores.reset_index(level=0, inplace=True)
    return cores
def load_GS_data(stain=stain):
    """
    loads the GS data for a stain and adds other columns relevant for analysis
    """
    coresGS = pd.read_excel("GS/GS_"+stain+".xlsx")
    coresGS = coresGS.rename(columns={
        'Core ID': 'coreID',
        '% Positive': 'expProp',
        'Intensity Score': 'expIntensity',
        'SQS': 'expSQS'
    })
    coresGS.insert(len(coresGS.columns),"expPropCategory",percentage_to_category(coresGS.expProp)[0])
    coresGS.insert(len(coresGS.columns),"expSQSadditive",coresGS.expIntensity + coresGS.expPropCategory)
    return coresGS
def core_dataframe_add_expert_scores(cores):
    """add expert scores and return the updated cores dataframe
    """
    # load in dataframe with columns
    # Core ID
    # % Positive
    # Intensity Score
    # SQS

    # add expert columns to dataframe
    cores = cores.merge(coresGS, how='left', on='coreID', sort=True)
    cores["hasExpert"] = ~np.isnan(cores.expSQS)
    return cores
def core_dataframe_add_corrected_SQS(cores):
    """
    This takes aggregateIntensityWeighted and aggregateProportionWeighted and corrects both of them individually through linear regression on expert scores.
    Then multiplies the two to come to a corrected SQS score. Note that by correcting p and i independently with both intercept and slope, you get non-linear changes in SQS.
    The cores for which we have expert data are corrected through 10-fold cross-validation; the data we do not have expert data for are corrected by estimating a single linear
    model for all the cores with expert data
    :param cores: pandas dataframe
    :return: cores with aggregateSQSCorrected, aggregatePropCorrected, aggregateIntensityCorrected added
    """
    # take out nans
    mask = np.array(~np.isnan(cores.expSQS).astype(bool))
    # prepare for regression
    Xprop = cores.aggregatePropWeighted[mask][:,np.newaxis]
    Yprop = cores.expProp[mask][:,np.newaxis]
    Xint = cores.aggregateIntensityWeighted[mask][:,np.newaxis]
    Yint = cores.expIntensity[mask][:,np.newaxis]

    nCores = len(cores.index)
    PredProp = np.ones((nCores,1))*np.nan
    PredInt = PredProp.copy()
    # prepare the sklearn function
    clf = linear_model.LinearRegression()
    if np.all(~np.isfinite(Xprop)) or (np.ma.size(Xprop,axis=1) == 0) or (np.ma.size(Xint, axis=1) == 0):  # if all values are NaN, or no columns
        cores["aggregatePropCorrected"] = np.nan
        cores["aggregateIntensityCorrected"] = np.nan
        cores["aggregatePropCorrectedCategory"] = np.nan
        cores["aggregateSQSCorrectedAdditive"] = np.nan
        cores["aggregateSQSCorrected"] = np.nan
    else:
        # prepare the sklearn pipeline. There's an imputer in case there's missing values - i.e. all subjects saying no cancer and
        # therefore not providing IHC scores
        p = pipeline.Pipeline([('imputer', Imputer(strategy='mean', missing_values='NaN')),('classifier', clf),])
        # use cross_val_predict to predict the values based on 10-fold cross-validation
        PredProp[mask] = cv.cross_val_predict(p,Xprop,Yprop,cv=10)
        p = pipeline.Pipeline([('imputer', Imputer(strategy='mean', missing_values='NaN')),('classifier', clf),])
        PredInt[mask] = cv.cross_val_predict(p,Xint,Yint,cv=10)
        # apply relationship to cores not done by experts only if they exist
        if np.sum(~mask)>0:
            # fit expert data
            p = pipeline.Pipeline([('imputer', Imputer(strategy='mean', missing_values='NaN')),('classifier', clf),])
            p.fit(X=Xprop,y=Yprop)
            PredProp[~mask] = p.predict(cores.aggregatePropWeighted[~mask][:,np.newaxis])
            # and again for intensity
            p = pipeline.Pipeline([('imputer', Imputer(strategy='mean', missing_values='NaN')),('classifier', clf),])
            p.fit(X=Xint,y=Yint)
            PredInt[~mask] = p.predict(cores.aggregateIntensityWeighted[~mask][:,np.newaxis])

        # filter any scores outside range
        np.clip(PredProp,0,100,PredProp)
        np.clip(PredInt,0,3,PredInt)
        # calculate product score
        PredSQS = PredProp*PredInt
        # add predicted/adjusted scores to cores
        cores["aggregatePropCorrected"] = PredProp
        cores["aggregateIntensityCorrected"] = PredInt
        cores["aggregatePropCorrectedCategory"] = percentage_to_category(cores["aggregatePropCorrected"])[0]
        cores["aggregateSQSCorrectedAdditive"] = cores["aggregateIntensityCorrected"] + cores["aggregatePropCorrectedCategory"]
        cores["aggregateSQSCorrected"] = PredSQS

    return cores
def core_dataframe_split_core_id(cores):
    """ Splits the 4 digit code and stain type into two columns called stain and coreID.
    """
    cores["stain"] = stain
    cores["stain"] = cores["stain"].astype(str)
    # get the number from the string
    cores["coreID"] = [int(re.search(r'\d+', x).group()) for foo,x in cores.core.iteritems()]
    return cores
def core_dataframe_write_to_excel(cores):
    """
    Writes two excel sheets: one with the whole core dataframe and one with only relevant columns for Anne
    :param cores:
    :return: None
    """
    # write entire dataframe
    cores.to_excel(excel_writer=("results\RtO_results_"+stain+"_"+aggregate+"_full.xlsx"))
    # write a clean version too
    c = cores.loc[:,["core",'aggregatePropWeighted','aggregatePropWeightedCategory','aggregateIntensityWeighted',\
                 'aggregateSQS','aggregateSQSadditive',"aggregatePropCorrected","aggregatePropCorrectedCategory",\
                     "aggregateIntensityCorrected",'aggregateSQSCorrected','aggregateSQSCorrectedAdditive','expProp','expIntensity','expSQS','expSQSadditive']]
    # rename columns for easier understanding
    c.rename(columns={
        'aggregatePropWeighted':'raw proportion',
        'aggregatePropWeightedCategory':'raw proportion category',
        'aggregateIntensityWeighted':'raw intensity',
        'aggregateSQS':'raw H-score',
        'aggregateSQSadditive':'raw Allred-like',

        "aggregatePropCorrected":'corrected proportion',
        "aggregatePropCorrectedCategory":'corrected proportion category',
        "aggregateIntensityCorrected":'corrected intensity',
        'aggregateSQSCorrected':'corrected H-score',
        'aggregateSQSCorrectedAdditive':'corrected Allred-like',

        'expProp':'expert proportion',
        "expPropCategory":'expert proportion category',
        'expIntensity':'expert intensity',
        'expSQS':'expert H-score',
        'expSQSadditive':'expert Allred-like'
    },inplace=True)
    # write separate excel file for this stain
    c.to_excel(excel_writer=("results\RtO_results_"+stain+"_"+aggregate+"_clean.xlsx"))
    # # write into a single aggregate excel doc with multiple sheets. Disabled for now because it's not useful
    # # and there's a problem with load_workbook seeing xlsx files as zip files if they are created through context menu in explorer
    # # copied from http://stackoverflow.com/questions/20219254/how-to-write-to-an-existing-excel-file-without-overwriting-data
    # aggregateFile = "results\RtO_results_clean.xlsx"
    # book = load_workbook(aggregateFile)
    # writer = pd.ExcelWriter(aggregateFile, engine='openpyxl')
    # writer.book = book
    # writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    # c.to_excel(excel_writer=writer,sheet_name=stain+aggregate)
    # writer.save()
def combine_all_stains_from_clean_excel_results():
    """Combines all the excel files *clean.xlsx into a single excel sheet, including patient IDs.

    On top of core_dataframe_write_to_excel(), will also add encrypted patient ID and stain

    This was requested by Anne Kiltie et al., email dated May 27 2016
    """
    # find all _clean.xlsx files from ignoring segments
    f = ['\\'.join(['results',fn]) for fn in os.listdir('results') if ('RtO_results_' in fn) and ('ignoring_segments_clean.xlsx' in fn)]
    # load all data into a pandas dataframe, starting with the first
    df = pd.read_excel(f[0])
    for fn in f[1:]:
        df_tmp = pd.read_excel(fn)
        # check all columns are the same
        assert all(df_tmp.columns == df.columns)
        # add new data as rows to existing dataframe
        df = df.append(df_tmp)
    # reset index
    df.reset_index(inplace=True, drop=True)
    # split core ID and stain
    df['stain'] = pd.Series([s.split()[1] for s in df.core])
    df['Core ID'] = [s.split()[0] for s in df.core]
    df['Core ID'] = df['Core ID'].astype(np.int)
    df.drop('core', axis=1, inplace=True)
    # now do a lookup to add patient ID
    lut = pd.read_excel('info/patient_core_LUT_encrypted.xlsx')
    df = df.merge(right=lut, how='left', left_on='Core ID', right_on='core', sort=True, copy=False).drop('core', axis=1, inplace=False)
    # write to excel
    fn = "results\cores_clean_results.xlsx"
    os.remove(fn) # delete existing file
    df.to_excel(excel_writer=(fn))


def core_dataframe_write_to_mongodb(cores):
    """ Inserts the records.

    :param cores: pandas dataframe
    :return:
    """
    con = MongoClient("localhost", 27017)
    db = con.results.cores
    # rename cores
    coresToWrite = cores.rename(columns = lambda x : aggregate + '.' + x)
    for _,core in coresToWrite.iterrows():
        # write basic info if core doesn't exist yet
        db.update_one({'stain': core[aggregate+'.stain'], "coreID": core[aggregate+'.coreID']}, {"$set":{'stain': core[aggregate+'.stain'], "coreID": core[aggregate+'.coreID']}}, upsert=True)
        # write in this method's data
        db.update_one({'stain': core[aggregate+'.stain'], "coreID": core[aggregate+'.coreID']}, {"$set": core.to_dict()})
def write_bootstrap_single_to_mongodb(dat, N):
    """ Takes a series of correlation values and writes them to mongodb

    :param dat: 1D array of values; see user_vs_expert_rho output
    :param N: number of users used to calculate aggregate
    :return:
    """
    db = dbConnection.results.bootstraps
    result = db.insert_one({
                'stain':stain,
                'nUsersPerSubject': int(N),
                'rhoProp':dat[0],
                'rhoIntensity':dat[1],
                'Hscore':dat[2],
                'Allred':dat[3],
                'qwkIntensity':dat[4],
                'last_modified':datetime.datetime.utcnow(),
                'database':rto_mongodb_utils.currentDB,
                'aggregate':aggregate
            }
        )
def get_core_ids(cln):
    """retrieves a list of cores expressed as objects for a given set of classifications
    Input is a dataframe with classifications containing a column called "core"
    """
    cores = list()
    for c in cln["core"]:
        if c not in cores:
            cores.append(c)
    return cores
def normalise_dataframe_by_ix(df,divideByColumn, columnsToDivide):
    """Returns dataframe with one set of columns divided by one of the columns in the dataframe.
    provide INDICES, not COLUMN NAMES. The columnsToDivide will be divided by divideByColumn
    """
    df2 = df.copy()
    # perform the division only for those columns requested
    return df2.iloc[:,columnsToDivide].div(df2.iloc[:,divideByColumn],axis="index")
def user_vs_expert_rho(cores):
    """Calculate Spearman rank correlation between expert and users and quadratic weighted kappa for intensity
    The function of scipy's spearmanr is undefined for nan so should use Panda's correlatin method instead, which only uses pairwise complete to compute r.
    :param cores: pandas dataframe
    :return: rhoProp,rhoIntensity,rhoSQS,rhoSQSadditive,qwkIntensity
    """
    rhoProp = cores['expProp'].corr(cores.aggregatePropCorrected,method='spearman')
    rhoIntensity = cores['expIntensity'].corr(cores.aggregateIntensityCorrected,method='spearman')
    rhoSQS = cores['expSQS'].corr(cores.aggregateSQSCorrected,method='spearman')
    rhoSQSadditive = cores['expSQSadditive'].corr(cores.aggregateSQSCorrectedAdditive,method='spearman')
    qwkIntensity = qwk.quadratic_weighted_kappa(cores.expIntensity, np.round(cores.aggregateIntensityCorrected), min_rating=0, max_rating=3)
    #There's a value consistently showing up in these calculations - must be something wrong. Always -0.378085328319
    if (rhoSQS < -0.378085328310) & (rhoSQS > -0.378085328320):
        cores.to_pickle('WEIRD_rhoSQS.pkl')
        raise Exception('check out the cores dataframe')

    return rhoProp,rhoIntensity,rhoSQS,rhoSQSadditive,qwkIntensity
def percentage_to_category(percentages):
    """Takes percentage stained and transforms to categories, useful for calculating pseudo-allred
    :param percentages: numpy array or Pandas series with percentage cancer. Can be None if only interested in the transformation
    :return: categ: same dimension as 'percentages', transformed to groups, float
    :return transform: pandas dataframe with a category and percentage column, indicating max percentage within that category
    """
    # set up dataframe with thresholds and categories. Percentages are inclusive. So if you're a 38, you'll be assigned the category which has the first value above 38 (e.g. 50)
    if stain in ('mre11','test mre11','p21','tip60','rad50','53bp1','ctip_nuclear','hdac2','nbs1','rpa','mre11c','dck_nuclear','mdm2', 'mre11new'):
        transform = pd.DataFrame(data=np.array([[0,1,2,3,4,5],[0,25,50,75,95,100],[0,12.5,37.5,62.5,85,97.5]]).T,columns=("category","percentage","middle_of_bin"))
    elif stain in ('p53','ki67'):
        transform = pd.DataFrame(data=np.array([[0,1,2,3,4,5],[0,10,25,50,75,100],[0,5,17.5,37.5,62.5,87.5]]).T,columns=("category","percentage","middle_of_bin"))
    elif stain in ('hdac4_membrane','ck5','ck20'):
        transform = pd.DataFrame(data=np.array([[0,1,2,3,4,5],[0,10,25,65,95,100],[0,5,17.5,45,80,97.5]]).T,columns=("category","percentage","middle_of_bin"))
    elif stain in ('ctip_cytoplasm','hdac4_cytoplasm','dck_cytoplasm'):
        raise Exception('unclear how to go from percentage to present/absent')
    else:
        raise Exception('transformation not specified for stain type')

    if np.all(percentages == None): # this would mean None was given
        return (None, transform)
    else: # this means non-None was given
        categ = np.zeros([len(percentages)])*np.nan
        for iP in range(len(percentages)):
            # might be a nan if user did not indicate cancer
            if not math.isnan(percentages[iP]):
                categ[iP] = min([x["category"] for _,x in transform.iterrows() if percentages[iP] <= x["percentage"]])
        return categ, transform
def category_to_percentage(categories):
    """NOTE: the response IDs in annotations[1]["a-2"] do not correspond to the categories expressed in percentage_to_category(),
    but are 1 higher. So when feeding them to this function, make sure to subtract 1 to make the range 0 to 5.
    If a category is -1 it is set a percentage = np.nan
    Takes a pd.series or np.array with staining categories and transforms them to percentages. Specifically, to percentage in the middle of the category
    Relies in percentage_to_category() to store the transformation from category to percentage
    :param categories: pandas series or np.array
    :return: np.array of same size as categories, but now with middle percentage of each category provided.
    """
    _,trans = percentage_to_category(None)
    # set index equal to categories
    foo = pd.DataFrame(categories, columns=["proportion"], dtype='float64')
    merged = foo.merge(trans,'left',left_on="proportion", right_on="category")
    # set -1 to
    merged[categories == -1] = np.nan
    return np.array(merged.middle_of_bin)
def plot_rho(rhoBoot):
    # set index as separate column
    toPlot = rhoBoot.copy()
    toPlot.reset_index(inplace=True)
    # toPlot.iloc[0:len(toPlot.index),1:].plot()
    toPlot.iloc[:,1:].plot()
    plt.xticks(np.arange(len(toPlot.index)),toPlot["index"])
    plt.xlabel("number of users per subject")
    plt.ylabel("Spearman r")
    plt.title("number of users/segment vs. accuracy")
    plt.draw()

def run_full_cores():
    """ Main function of this script.
    Will run a single pass through analysis, collecting classifications, forming dataframe, writing to excel and mongodb
    :return: cores
    """
    if aggregate == 'segment_aggregation':
        # this code should only be ran when including all users
        assert numberOfUsersPerSubject[0]==0
        assert len(numberOfUsersPerSubject)==1
        assert samplesPerNumberOfUsers==1
        # check if dataframe with classifications exists; if not, run over each classification and store its properties in a pandas dataframe. If it does exist, load it.
        if os.path.isfile(classificationsDataframeFn) == False:
            cln = classifications_dataframe_fill(numberOfUsersPerSubject=0,skipNonExpertClassifications=False)
            classifications_dataframe_save(cln)
        else:  # dataframe available, load instead of fill
            cln = classifications_dataframe_load(fn=classificationsDataframeFn)
        # add columns to cln that indicates, for each subject, what the aggregate is of the multiple columns
        cln = cln_add_columns_aggregating_stain(cln)
        # aggregate data from multiple subjects into a single score for each core
        cores = core_dataframe_fill(cln)
    elif aggregate == 'ignoring_segments':
        assert numberOfClassificationsPerCore[0] == 0
        cln = classifications_dataframe_fill_individual_classifications(skipNonExpertClassifications=False)
        cores = cores_dataframe_fill_from_individual_classifications(cln=cln, nCl=numberOfClassificationsPerCore[0])
    cores = core_dataframe_split_core_id(cores)
    cores = core_dataframe_add_expert_scores(cores)
    cores = core_dataframe_add_corrected_SQS(cores)
    core_dataframe_write_to_mongodb(cores)
    core_dataframe_write_to_excel(cores)
    combine_all_stains_from_clean_excel_results()
    return (cores,cln)
def run_bootstrap_rho_segment_aggregation():
    dbBootstraps = dbConnection.results.bootstraps
    ix = 0
    for N in numberOfUsersPerSubject:
        rho = np.zeros([samplesPerNumberOfUsers,5])
        t = time.time()
        bootstrapCount = dbBootstraps.find({'stain': stain, 'nUsersPerSubject': N, 'aggregate': aggregate}).count()
        # either some samples remain, or 0 remain (can't be negative)
        toDo = max(0, samplesPerNumberOfUsers - bootstrapCount)
        print("Already %d bootstraps in database, doing %d more" % (bootstrapCount, toDo))
        for iB in range(toDo):
            cln = classifications_dataframe_fill(numberOfUsersPerSubject=N,skipNonExpertClassifications=True)
            cln = cln_add_columns_aggregating_stain(cln)
            cores = core_dataframe_fill(cln)
            cores = core_dataframe_split_core_id(cores)
            cores = core_dataframe_add_expert_scores(cores)
            cores = core_dataframe_add_corrected_SQS(cores)
            rho = user_vs_expert_rho(cores)
            print("including %d users; completed %d out of %d. Elapsed time for this bootstrap: %d seconds" % (
                N, iB+1, toDo, np.round(time.time()-t)))
            # write this iteration to mongodb
            write_bootstrap_single_to_mongodb(rho,N)
        ix += 1
def run_bootstrap_rho_ignoring_segments():
    ix = 0
    clnAll = classifications_dataframe_fill_individual_classifications(skipNonExpertClassifications=True)
    dbBootstraps = dbConnection.results.bootstraps
    for N in numberOfClassificationsPerCore:
        t = time.time()
        # check how many bootstrap entries there are
        bootstrapCount = dbBootstraps.find({'stain': stain, 'nUsersPerSubject': int(N), 'aggregate': aggregate}).count()
        # either some samples remain, or 0 remain (can't be negative)
        toDo = max(0, samplesPerNumberOfUsers - bootstrapCount)
        print("Already %d bootstraps in database, doing %d more" % (bootstrapCount, toDo))
        for iB in range(toDo):
            cores = cores_dataframe_fill_from_individual_classifications(cln=clnAll, nCl=N)
            cores = core_dataframe_split_core_id(cores)
            cores = core_dataframe_add_expert_scores(cores)
            cores = core_dataframe_add_corrected_SQS(cores)
            rho = user_vs_expert_rho(cores)
            print("including %d users; completed %d out of %d. Elapsed time for this bootstrap: %d seconds" % (
                N, iB + 1, toDo, np.round(time.time() - t)))
            # write this iteration to mongodb
            write_bootstrap_single_to_mongodb(rho,N)
            # check how many bootstrap entries there are
            bootstrapCount = dbBootstraps.find({'stain': stain, 'nUsersPerSubject': int(N), 'aggregate': aggregate}).count()



########### FUNCTION EXECUTION
def main():
    if not bootstrap:
        cores, cln = run_full_cores() # will also generate the .pkl file with classifications if it doesn't exist
    elif (aggregate == 'segment_aggregation') & bootstrap:
        run_bootstrap_rho_segment_aggregation() # requires run_full_cores to have run; requires rto_mongodb_utils.
    elif (aggregate == 'ignoring_segments') & bootstrap:
        run_bootstrap_rho_ignoring_segments()

    # plot_weighted_vs_unweighted_stain(cores)
    # plot_user_vs_expert(cores)

    # close the connection to local mongoDB
    pymongo_connection_close()
    print("Finished user_aggregation.py")

# load GS data
coresGS = load_GS_data(stain)

# only execute code if the code is being ran on its own
if __name__ == "__main__":
    # get these so they're available in each function's namespace without having to pass
    subjectsCollection, classifCollection, dbConnection = pymongo_connection_open()
    main()
